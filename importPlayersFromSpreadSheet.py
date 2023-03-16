import pandas as pd
import openpyxl
from player import Player



def extract_data_using_panda(path_link):
    df = pd.read_excel(path_link)
    player_list = []
    for index, row in df.iterrows():
        player_name = (row[0])
        player_mail = (row[1])
        player_participate = (row[2])
        player_lastpick = (row[3])
        if player_participate == "Y":
            player_list.append(Player(player_name, player_mail))
            player_list[len(player_list) - 1].addRestriction(player_lastpick)
    return player_list


def extract_data_using_openpyxl(path_link):
    inv_file = openpyxl.load_workbook(path_link)
    sheet = inv_file.active
    player_list = []

    for file_row in range(2, sheet.max_row + 1):
        player_name = sheet.cell(file_row, 1).value
        player_email = sheet.cell(file_row, 2).value
        player_participate = sheet.cell(file_row, 3).value
        player_lastpick = sheet.cell(file_row, 4).value


        if player_name is not None and player_participate == "Y":
            player_list.append(Player(player_name, player_email))
            player_list[len(player_list)-1].addRestriction(player_lastpick)
    return player_list


def update_lastpick_in_file(path_link, player_dictionnary):
    inv_file = openpyxl.load_workbook(path_link)
    sheet = inv_file.active
    player_list = []

    for file_row in range(2, sheet.max_row + 1):
        player_name = sheet.cell(file_row, 1).value
        player_participate = sheet.cell(file_row, 3).value
        player_lastpick = sheet.cell(file_row, 4)

        if player_name is not None:

            if player_participate == "Y":
                player_object = player_dictionnary[player_name]
                player_lastpick.value = player_object.secretPick[0]
            else:
                player_lastpick.value = ""

    inv_file.save("newSecretHannouka.xlsx")
    #os.remove("SecretHannouka.xlsx")


#extract_data_using_openpyxl("/Users/samueltenoudji-cohen/Desktop/SecretHanoukka.xlsx")




